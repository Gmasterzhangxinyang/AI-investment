from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_sheet = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
    format_workbook(path)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="172033")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 12
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value) + 2, 42))
            ws.column_dimensions[col_letter].width = min(max_len, 34)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(path)

